"""PART 1
==============================================================================
Excel GST Automation
Author : Supriya Bhowmik + ChatGPT
Version : 1.0

This software will automate

PHASE-1
    • Create Sheet 5
    • Create Sheet 18
    • Copy GST rows
    • Preserve formatting
    • Add GST summary rows

PHASE-2
    • Create Temp_5

PHASE-3
    • Create Temp_18
==============================================================================

"""

import copy
import logging
from pathlib import Path
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font
import tkinter as tk
from tkinter import filedialog
###############################################################################
# CONFIGURATION
###############################################################################

SOURCE_SHEET = "Bill Details"
ITEM_SHEET = "ITEM"

SHEET_5 = "5"
SHEET_18 = "18"

TEMP_5 = "Temp_5"
TEMP_18 = "Temp_18"

GST5 = 0.05
GST18 = 0.18

RED = "FF0000"

TODAY = datetime.today()

###############################################################################
# LOGGING
###############################################################################

logging.basicConfig(
    level=logging.INFO,
    format="%(message)s"
)

###############################################################################
# WORKBOOK
###############################################################################

class GSTWorkbook:

    def __init__(self, filename):

        self.filename = filename

        logging.info("Loading workbook...")

        self.wb = load_workbook(filename)

        self.bill = self.wb[SOURCE_SHEET]

        self.item = self.wb[ITEM_SHEET]

        self.header_row = 5

        self.columns = {}

        self.farmer_name = self.bill["C2"].value

        self.map_columns()

    ###########################################################################

    def map_columns(self):

        """
        Read all headers from Row-5
        """

        for cell in self.bill[self.header_row]:

            if cell.value is None:
                continue

            self.columns[
                str(cell.value).strip().lower()
            ] = cell.column

        logging.info("Columns detected")

        for k, v in self.columns.items():

            logging.info(f"{k:<20} -> {v}")

    ###########################################################################

    def column(self, name):

        name = name.lower()

        if name not in self.columns:

            raise Exception(
                f"Column '{name}' not found."
            )

        return self.columns[name]

###############################################################################
# STYLE COPYING
###############################################################################

def copy_style(source, target):

    if source.has_style:

        target.font = copy.copy(source.font)

        target.fill = copy.copy(source.fill)

        target.border = copy.copy(source.border)

        target.alignment = copy.copy(source.alignment)

        target.number_format = copy.copy(source.number_format)

        target.protection = copy.copy(source.protection)

###############################################################################
# COPY ROW
###############################################################################

def copy_row(source_ws, source_row, target_ws, target_row):

    for cell in source_ws[source_row]:

        new_cell = target_ws.cell(
            row=target_row,
            column=cell.column
        )

        new_cell.value = cell.value

        copy_style(
            cell,
            new_cell
        )

    if source_row in source_ws.row_dimensions:

        target_ws.row_dimensions[target_row].height = \
            source_ws.row_dimensions[source_row].height

###############################################################################
# COPY COLUMN WIDTHS
###############################################################################

def copy_column_widths(source_ws, target_ws):

    for key, value in source_ws.column_dimensions.items():

        target_ws.column_dimensions[key].width = value.width

###############################################################################
# CREATE OUTPUT SHEETS
###############################################################################

def prepare_sheet(wb, name):

    if name in wb.sheetnames:

        del wb[name]

    return wb.create_sheet(name)

###############################################################################
# FIND LAST ROW
###############################################################################

def last_row(ws):

    r = ws.max_row

    while r > 1:

        ok = False

        for c in ws[r]:

            if c.value not in ("", None):

                ok = True

                break

        if ok:

            return r

        r -= 1

    return 1

###############################################################################
# RED BOLD
###############################################################################

def red_bold(cell):

    cell.font = Font(
        name=cell.font.name,
        size=cell.font.size,
        bold=True,
        color=RED
    )

###############################################################################
# GST CHECK
###############################################################################

def is_gst_5(value):

    try:

        return round(float(value),2)==0.05

    except:

        return False


def is_gst_18(value):

    try:

        return round(float(value),2)==0.18

    except:

        return False


###############################################################################
# PART-1 COMPLETE
###############################################################################

print("\n")
print("="*70)
print("PART-1 LOADED SUCCESSFULLY")
print("="*70)
print("\n")

###############################################################################
# COPY COMPLETE SHEET LAYOUT PART-2 START
###############################################################################

def copy_sheet_layout(source_ws, target_ws):
    """
    Copy complete worksheet layout.
    """

    # Copy column widths
    for col_letter, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dim.width
        target_ws.column_dimensions[col_letter].hidden = dim.hidden

    # Copy row heights
    for row_no, dim in source_ws.row_dimensions.items():
        target_ws.row_dimensions[row_no].height = dim.height
        target_ws.row_dimensions[row_no].hidden = dim.hidden

    # Freeze panes
    target_ws.freeze_panes = source_ws.freeze_panes

    # Sheet View
    target_ws.sheet_view.zoomScale = source_ws.sheet_view.zoomScale
    target_ws.sheet_view.showGridLines = \
        source_ws.sheet_view.showGridLines

    # Merged Cells
    for merged in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged))


###############################################################################
# COPY COMPLETE ROW
###############################################################################

def copy_complete_row(source_ws,
                      source_row,
                      target_ws,
                      target_row):

    for cell in source_ws[source_row]:

        # ---------------------------------------------------------------
        # MergedCell objects are read-only.
        # Only the top-left cell of a merged range can contain a value.
        # ---------------------------------------------------------------
        if isinstance(cell, MergedCell):

            continue

        new_cell = target_ws.cell(
            row=target_row,
            column=cell.column
        )

        # Copy value or formula
        new_cell.value = cell.value

        # Copy formatting
        copy_style(
            cell,
            new_cell
        )

    # Preserve row height
    if source_row in source_ws.row_dimensions:

        target_ws.row_dimensions[target_row].height = \
            source_ws.row_dimensions[source_row].height


###############################################################################
# CREATE OUTPUT SHEETS
###############################################################################

def create_output_sheets(book):

    logging.info("Creating Sheet 5...")

    if SHEET_5 in book.wb.sheetnames:
        del book.wb[SHEET_5]

    sheet5 = book.wb.create_sheet(SHEET_5)

    logging.info("Creating Sheet 18...")

    if SHEET_18 in book.wb.sheetnames:
        del book.wb[SHEET_18]

    sheet18 = book.wb.create_sheet(SHEET_18)

    copy_sheet_layout(
        book.bill,
        sheet5
    )

    copy_sheet_layout(
        book.bill,
        sheet18
    )

    return sheet5, sheet18


###############################################################################
# COPY HEADER SECTION
###############################################################################

def copy_header(book,
                target_ws):

    for row in range(1, book.header_row + 1):

        copy_complete_row(
            book.bill,
            row,
            target_ws,
            row
        )


###############################################################################
# SPLIT GST ROWS
###############################################################################

def split_gst(book,
              sheet5,
              sheet18):

    logging.info("Splitting GST Rows...")

    gst_col = book.column("gst rate")

    target5 = book.header_row + 1

    target18 = book.header_row + 1

    last = last_row(book.bill)

    for row in range(book.header_row + 1,
                     last + 1):

        gst = book.bill.cell(
            row=row,
            column=gst_col
        ).value

        if is_gst_5(gst):

            copy_complete_row(
                book.bill,
                row,
                sheet5,
                target5
            )

            target5 += 1

        elif is_gst_18(gst):

            copy_complete_row(
                book.bill,
                row,
                sheet18,
                target18
            )

            target18 += 1

    logging.info("GST Split Completed")

    return target5, target18


#####################################################
# SAVE WORKBOOK
#####################################################

def save(book):

    file = Path(book.filename)

    output = file.parent / \
        ("Revised " + file.name)

    #####################################################
    # REQUIRED SHEET ORDER
    #####################################################

    required_order = [
        "5",
        "18",
        "Temp_5",
        "Temp_18",
        "Bill Details",
        "ITEM"
    ]

    #####################################################
    # REORDER EXISTING WORKSHEETS
    #####################################################

    ordered_sheets = []

    for sheet_name in required_order:

        if sheet_name in book.wb.sheetnames:

            ordered_sheets.append(
                book.wb[sheet_name]
            )

    #####################################################
    # APPLY THE REQUIRED WORKSHEET ORDER
    #####################################################

    book.wb._sheets = ordered_sheets

    #####################################################
    # SAVE WORKBOOK
    #####################################################

    book.wb.save(output)

    logging.info("Workbook Saved")

    logging.info(output)

###############################################################################
# CALCULATE BILL VALUES
###############################################################################

def calculate_bill_values(ws, header_row=5):

    """
    Calculate:

    G = Item Vaue
    I = CGST
    J = SGST
    K = Final Amount

    This function works on the copied GST sheets.
    """

    # Fixed columns from your workbook
    COL_ITEMS = 2
    COL_QUANTITY = 4
    COL_SALE_RATE = 6
    COL_ITEM_VALUE = 7
    COL_GST_RATE = 8
    COL_CGST = 9
    COL_SGST = 10
    COL_FINAL_AMOUNT = 11

    for row in range(header_row + 1, ws.max_row + 1):

        item = ws.cell(row=row, column=COL_ITEMS).value

        if item is None or str(item).strip() == "":
            continue

        item_text = str(item).strip().upper()

        # ---------------------------------------------------------------
        # Read GST Rate
        # ---------------------------------------------------------------
        gst_rate = ws.cell(
            row=row,
            column=COL_GST_RATE
        ).value

        try:
            gst_rate = float(gst_rate)

        except (TypeError, ValueError):

            # If GST rate is missing, do not calculate tax
            continue

        # ---------------------------------------------------------------
        # Special handling for FREIGHT row
        # ---------------------------------------------------------------
        if "FREIGHT CHARGES FOR EC SHED" in item_text:

    # The source Final Amount is the GST-inclusive amount.
    # Reverse GST to calculate the taxable Item Value.

            source_final_amount = ws.cell(
                row=row,
                column=COL_FINAL_AMOUNT
            ).value

            try:

                gross_amount = float(
                    source_final_amount
                )

            except (TypeError, ValueError):

                continue

    # Reverse GST:
    #
    # Taxable Value =
    # Gross Amount / (1 + GST Rate)
    #
    # Example:
    # 10949.94 / 1.18 = 9279.61

            item_value = round(
                gross_amount / (1 + gst_rate),
                2
            )

    # CGST and SGST are half of total GST

            cgst = round(
                item_value * gst_rate / 2,
                2
            )

            sgst = round(
                item_value * gst_rate / 2,
                2
            )

    # Preserve the original GST-inclusive amount

            final_amount = round(
                gross_amount,
                2
            )

        else:

            quantity = ws.cell(
                row=row,
                column=COL_QUANTITY
            ).value

            sale_rate = ws.cell(
                row=row,
                column=COL_SALE_RATE
            ).value

            try:

                quantity = float(quantity)

                sale_rate = float(sale_rate)

                item_value = round(
                    quantity * sale_rate,
                    2
                )

                cgst = round(
                    item_value * gst_rate / 2,
                    2
                )

                sgst = round(
                    item_value * gst_rate / 2,
                    2
                )

                final_amount = round(
                    item_value + cgst + sgst,
                    2
                )

            except (TypeError, ValueError):

                continue

        # ---------------------------------------------------------------
        # Calculate GST
        # ---------------------------------------------------------------

        #cgst = item_value * gst_rate / 2

        #sgst = item_value * gst_rate / 2

        #final_amount = item_value + cgst + sgst

        # ---------------------------------------------------------------
        # Write calculated values
        # ---------------------------------------------------------------

        ws.cell(
            row=row,
            column=COL_ITEM_VALUE
        ).value = round(item_value, 2)

        ws.cell(
            row=row,
            column=COL_CGST
        ).value = round(cgst, 2)

        ws.cell(
            row=row,
            column=COL_SGST
        ).value = round(sgst, 2)

        ws.cell(
            row=row,
            column=COL_FINAL_AMOUNT
        ).value = round(final_amount, 2)

        # ---------------------------------------------------------------
        # Apply number format
        # ---------------------------------------------------------------

        ws.cell(
            row=row,
            column=COL_ITEM_VALUE
        ).number_format = "0.00"

        ws.cell(
            row=row,
            column=COL_CGST
        ).number_format = "0.00"

        ws.cell(
            row=row,
            column=COL_SGST
        ).number_format = "0.00"

        ws.cell(
            row=row,
            column=COL_FINAL_AMOUNT
        ).number_format = "0.00"

###############################################################################
# RENUMBER SERIAL COLUMN
###############################################################################

def renumber_sl(ws, header_row=5):

    """
    Renumber Sl column chronologically starting from 1.
    """

    sl_col = 1
    item_col = 2

    serial = 1

    for row in range(header_row + 1, ws.max_row + 1):

        item = ws.cell(
            row=row,
            column=item_col
        ).value

        if item is None or str(item).strip() == "":
            continue

        ws.cell(
            row=row,
            column=sl_col
        ).value = serial

        serial += 1


###############################################################################
# ADD GST SUMMARY ROWS
###############################################################################

def add_gst_summary_rows(
        ws,
        header_row,
        cgst_description,
        sgst_description):

    """
    Add two summary rows at the end of a GST sheet.

    CGST column = I
    SGST column = J
    Items column = B
    Sl column = A
    """

    ITEM_COL = 2
    CGST_COL = 9
    SGST_COL = 10
    SL_COL = 1

    # ---------------------------------------------------------------
    # Find the last existing data row
    # ---------------------------------------------------------------

    last_data = last_row(ws)

    cgst_row = last_data + 1

    sgst_row = last_data + 2

    # ---------------------------------------------------------------
    # Copy style from the previous data row
    # ---------------------------------------------------------------

    for col in range(1, ws.max_column + 1):

        source_cell = ws.cell(
            row=last_data,
            column=col
        )

        cgst_cell = ws.cell(
            row=cgst_row,
            column=col
        )

        sgst_cell = ws.cell(
            row=sgst_row,
            column=col
        )

        copy_style(
            source_cell,
            cgst_cell
        )

        copy_style(
            source_cell,
            sgst_cell
        )

    # ---------------------------------------------------------------
    # Sl numbers
    # ---------------------------------------------------------------

    previous_sl = ws.cell(
        row=last_data,
        column=SL_COL
    ).value

    try:

        next_sl = int(previous_sl) + 1

    except (TypeError, ValueError):

        next_sl = last_data - header_row + 1

    ws.cell(
        row=cgst_row,
        column=SL_COL
    ).value = next_sl

    ws.cell(
        row=sgst_row,
        column=SL_COL
    ).value = next_sl + 1

    # ---------------------------------------------------------------
    # Description
    # ---------------------------------------------------------------

    ws.cell(
        row=cgst_row,
        column=ITEM_COL
    ).value = cgst_description

    ws.cell(
        row=sgst_row,
        column=ITEM_COL
    ).value = sgst_description

    # ---------------------------------------------------------------
    # Dynamic SUM formulas
    # ---------------------------------------------------------------

    first_data_row = header_row + 1

    last_original_data_row = last_data

    ws.cell(
        row=cgst_row,
        column=CGST_COL
    ).value = (
        f"=SUM(I{first_data_row}:I{last_original_data_row})"
    )

    ws.cell(
        row=sgst_row,
        column=SGST_COL
    ).value = (
        f"=SUM(J{first_data_row}:J{last_original_data_row})"
    )

    # ---------------------------------------------------------------
    # Format totals
    # ---------------------------------------------------------------

    cgst_total_cell = ws.cell(
        row=cgst_row,
        column=CGST_COL
    )

    sgst_total_cell = ws.cell(
        row=sgst_row,
        column=SGST_COL
    )

    red_bold(cgst_total_cell)

    red_bold(sgst_total_cell)

    cgst_total_cell.number_format = "0.00"

    sgst_total_cell.number_format = "0.00"

    # ---------------------------------------------------------------
    # Ensure blank cells in summary rows remain blank
    # ---------------------------------------------------------------

    for col in range(1, ws.max_column + 1):

        if col not in (
            SL_COL,
            ITEM_COL,
            CGST_COL,
            SGST_COL
        ):

            ws.cell(
                row=cgst_row,
                column=col
            ).value = None

            ws.cell(
                row=sgst_row,
                column=col
            ).value = None

    logging.info(
        f"GST summary rows added to '{ws.title}'"
    )

###############################################################################
# COMPLETE PHASE 1 SHEET PROCESSING
###############################################################################

def complete_gst_sheet(
        ws,
        header_row,
        cgst_description,
        sgst_description):

    logging.info(
        f"Renumbering Sheet '{ws.title}'..."
    )

    renumber_sl(
        ws,
        header_row
    )

    logging.info(
        f"Adding GST totals to Sheet '{ws.title}'..."
    )

    add_gst_summary_rows(
        ws,
        header_row,
        cgst_description,
        sgst_description
    )

###############################################################################
# TEMP_5 HEADERS
###############################################################################

TEMP_HEADERS = [
    "Ctr",
    "Voucher Date",
    "Voucher Type Name",
    "Voucher Number",
    "Ledger Name",
    "Ledger Amount",
    "Ledger Amount Dr/Cr",
    "Item Name",
    "Item Allocations - Godown Name",
    "Actual Quantity",
    "Billed Quantity",
    "Quantity UOM",
    "Item Rate",
    "Item Rate per",
    "Item Amount",
    "Change Mode"
]

###############################################################################
# CREATE TEMP_5 SHEET
###############################################################################
from datetime import date
def create_temp_5(book, sheet5):

    logging.info("Creating Temp_5...")

    # Remove existing Temp_5 if present
    if TEMP_5 in book.wb.sheetnames:
        del book.wb[TEMP_5]

    temp5 = book.wb.create_sheet(TEMP_5)

    # ---------------------------------------------------------------
    # Write headers
    # ---------------------------------------------------------------

    for col, header in enumerate(TEMP_HEADERS, start=1):

        cell = temp5.cell(
            row=1,
            column=col
        )

        cell.value = header

        # Header formatting
        cell.font = copy.copy(
            sheet5.cell(
                row=book.header_row,
                column=1
            ).font
        )

        cell.fill = copy.copy(
            sheet5.cell(
                row=book.header_row,
                column=1
            ).fill
        )

        cell.border = copy.copy(
            sheet5.cell(
                row=book.header_row,
                column=1
            ).border
        )

        cell.alignment = copy.copy(
            sheet5.cell(
                row=book.header_row,
                column=1
            ).alignment
        )

    # ---------------------------------------------------------------
    # Sheet 5 Sl rows + one Farmer Name row
    # ---------------------------------------------------------------

    #total_rows = sheet5.max_row - book.header_row
    total_rows = sheet5.max_row - book.header_row + 1

    # ---------------------------------------------------------------
    # Create rows
    # ---------------------------------------------------------------

    for row in range(2, total_rows + 2):

        ctr = row - 2

        # Ctr
        temp5.cell(
            row=row,
            column=1
        ).value = ctr

        # Voucher Date
        temp5.cell(
            row=row,
            column=2
        ).value = date.today()

        temp5.cell(
            row=row,
            column=2
        ).number_format = "DD/MM/YY"

        # Voucher Type Name
        temp5.cell(
            row=row,
            column=3
        ).value = "Sale_EC Shed"

        # Voucher Number
        temp5.cell(
            row=row,
            column=4
        ).value = (
            f"HHF/E/"
            f"{TODAY.month:02d}/1/2627"
        )

        # Ledger Amount Dr/Cr
        temp5.cell(
            row=row,
            column=7
        ).value = "Dr" if row == 2 else "Cr"

        # Item Allocations - Godown Name
        if row != 2:

            temp5.cell(
                row=row,
                column=9
            ).value = "Dankuni"

        # Change Mode
        temp5.cell(
            row=row,
            column=16
        ).value = "Item Invoice"

    logging.info(
        "Temp_5 created successfully"
    )

    return temp5

###############################################################################
# CREATE TEMP_18 SHEET
###############################################################################
from datetime import date
def create_temp_18(book, sheet18):

    logging.info("Creating Temp_18...")

    # Remove existing Temp_18 if present
    if TEMP_18 in book.wb.sheetnames:
        del book.wb[TEMP_18]

    temp18 = book.wb.create_sheet(TEMP_18)

    # -------------------------------------------------------------------------
    # Write headers
    # -------------------------------------------------------------------------

    for col, header in enumerate(TEMP_HEADERS, start=1):

        cell = temp18.cell(
            row=1,
            column=col
        )

        cell.value = header

        # Copy header formatting from Sheet 18
        cell.font = copy.copy(
            sheet18.cell(
                row=book.header_row,
                column=1
            ).font
        )

        cell.fill = copy.copy(
            sheet18.cell(
                row=book.header_row,
                column=1
            ).fill
        )

        cell.border = copy.copy(
            sheet18.cell(
                row=book.header_row,
                column=1
            ).border
        )

        cell.alignment = copy.copy(
            sheet18.cell(
                row=book.header_row,
                column=1
            ).alignment
        )

    # -------------------------------------------------------------------------
    # IMPORTANT MAPPING
    #
    # Sheet 18:
    #
    # Sl 1 to Sl N
    #
    # Temp_18:
    #
    # Ctr 0  = Farmer Name
    # Ctr 1  = Sheet 18 Sl 1
    # Ctr 2  = Sheet 18 Sl 2
    # ...
    # Ctr N  = Sheet 18 Sl N
    #
    # Therefore:
    #
    # Temp_18 rows = Sheet 18 Sl rows + 1
    # -------------------------------------------------------------------------

    total_rows = sheet18.max_row - book.header_row + 1

    # -------------------------------------------------------------------------
    # Create data rows
    # -------------------------------------------------------------------------

    for row in range(2, total_rows + 2):

        # Excel row 2 -> Ctr 0
        # Excel row 3 -> Ctr 1
        # Excel row 4 -> Ctr 2
        # ...
        # Ctr N -> Sheet 18 Sl N

        ctr = row - 2

        # Ctr
        temp18.cell(
            row=row,
            column=1
        ).value = ctr

        # Voucher Date
        temp18.cell(
            row=row,
            column=2
        ).value = date.today()

        temp18.cell(
            row=row,
            column=2
        ).number_format = "DD/MM/YY"

        # Voucher Type Name
        temp18.cell(
            row=row,
            column=3
        ).value = "Sale_EC Shed"

        # Voucher Number
        temp18.cell(
            row=row,
            column=4
        ).value = (
            f"HHF/E/"
            f"{TODAY.month:02d}/2/2627"
        )

        # Dr on Farmer Name row
        # Cr on all subsequent rows
        temp18.cell(
            row=row,
            column=7
        ).value = "Dr" if ctr == 0 else "Cr"

        # Godown Name
        if ctr != 0:

            temp18.cell(
                row=row,
                column=9
            ).value = "Dankuni"

        # Change Mode
        temp18.cell(
            row=row,
            column=16
        ).value = "Item Invoice"

    logging.info(
        "Temp_18 created successfully"
    )

    return temp18

###############################################################################
# FIND FARMER LEDGER NAME
###############################################################################

def get_farmer_ledger_name(book):

    """
    Find the description:
        Farmer Name (As per Tally Ledger)

    and return the value from the immediate right-side cell.
    """

    target_text = "Farmer Name (As per Tally Ledger)"

    for row in book.bill.iter_rows():

        for cell in row:

            if str(cell.value).strip() == target_text:

                right_cell = book.bill.cell(
                    row=cell.row,
                    column=cell.column + 1
                )

                return right_cell.value

    raise ValueError(
        "Could not find 'Farmer Name (As per Tally Ledger)' "
        "in Bill Details."
    )

###############################################################################
# POPULATE TEMP_5
###############################################################################

def populate_temp_5(book, temp5, sheet5):

    logging.info("Populating Temp_5 formulas...")

    HEADER_ROW = book.header_row

    # -------------------------------------------------------------------------
    # Sheet 5 structure
    #
    # Example:
    #
    # Header row       = 5
    # Sl 1             = row 6
    # Sl 2             = row 7
    # ...
    # Sl 12            = CGST
    # Sl 13            = SGST
    #
    # -------------------------------------------------------------------------

    first_sheet5_data_row = HEADER_ROW + 1

    last_sheet5_row = sheet5.max_row

    # Number of rows represented by Sl in Sheet 5
    #
    # Example:
    # Sheet 5 rows 6 to 18
    # 18 - 5 = 13 rows
    #
    total_sl_rows = last_sheet5_row - HEADER_ROW

    # -------------------------------------------------------------------------
    # Temp_5 must contain:
    #
    # Ctr 0  = Farmer Name
    # Ctr 1  = Sheet 5 Sl 1
    # Ctr 2  = Sheet 5 Sl 2
    # ...
    # Ctr 13 = Sheet 5 Sl 13
    #
    # Therefore:
    #
    # Number of Temp_5 data rows =
    # Number of Sheet 5 Sl rows + 1 Farmer Name row
    #
    # Example:
    # 13 Sheet 5 rows + 1 Farmer Name row = 14 Temp_5 rows
    # -------------------------------------------------------------------------

    total_temp5_rows = total_sl_rows + 1
    #total_rows = sheet5.max_row - book.header_row + 1

    # -------------------------------------------------------------------------
    # GST summary Ctr values
    #
    # Sheet 5:
    #
    # Sl 12 = CGST
    # Sl 13 = SGST
    #
    # Therefore:
    #
    # Ctr 12 = CGST
    # Ctr 13 = SGST
    # -------------------------------------------------------------------------

    cgst_ctr = total_sl_rows - 1

    sgst_ctr = total_sl_rows

    # -------------------------------------------------------------------------
    # Find farmer ledger name
    # -------------------------------------------------------------------------

    farmer_ledger_name = get_farmer_ledger_name(book)

    # -------------------------------------------------------------------------
    # Process Temp_5 rows
    # -------------------------------------------------------------------------

    for temp_row in range(2, total_temp5_rows + 2):

        # ---------------------------------------------------------------------
        # Ctr calculation
        #
        # Excel row 2 → Ctr 0
        # Excel row 3 → Ctr 1
        # Excel row 4 → Ctr 2
        # ...
        # ---------------------------------------------------------------------

        ctr = temp_row - 2

        # =====================================================================
        # CTR 0 — FARMER NAME ROW
        # =====================================================================

        if ctr == 0:

            # Ledger Name
            temp5.cell(
                row=temp_row,
                column=5
            ).value = farmer_ledger_name

            # Ledger Amount
            #
            # Sum of Final Amount column K in Sheet 5
            #

            temp5.cell(
                row=temp_row,
                column=6
            ).value = (
                f"=SUM('5'!K{first_sheet5_data_row}:"
                f"K{last_sheet5_row})"
            )

        # =====================================================================
        # CTR 12 — CGST SUMMARY ROW
        # =====================================================================

        elif ctr == cgst_ctr:

            # Ledger Name
            #
            # Ctr 12 → Sheet 5 Sl 12
            #

            temp5.cell(
                row=temp_row,
                column=5
            ).value = (
                f"=INDEX('5'!B:B,"
                f"MATCH(A{temp_row},'5'!A:A,0))"
            )

            # Ledger Amount
            #
            # Ctr 12 → Sheet 5 Sl 12 → CGST column I
            #

            temp5.cell(
                row=temp_row,
                column=6
            ).value = (
                f"=INDEX('5'!I:I,"
                f"MATCH(A{temp_row},'5'!A:A,0))"
            )

        # =====================================================================
        # CTR 13 — SGST SUMMARY ROW
        # =====================================================================

        elif ctr == sgst_ctr:

            # Ledger Name
            #
            # Ctr 13 → Sheet 5 Sl 13
            #

            temp5.cell(
                row=temp_row,
                column=5
            ).value = (
                f"=INDEX('5'!B:B,"
                f"MATCH(A{temp_row},'5'!A:A,0))"
            )

            # Ledger Amount
            #
            # Ctr 13 → Sheet 5 Sl 13 → SGST column J
            #

            temp5.cell(
                row=temp_row,
                column=6
            ).value = (
                f"=INDEX('5'!J:J,"
                f"MATCH(A{temp_row},'5'!A:A,0))"
            )

        # =====================================================================
        # NORMAL ITEM ROWS
        # =====================================================================

        else:

            # -----------------------------------------------------------------
            # Ledger Name
            # -----------------------------------------------------------------

            temp5.cell(
                row=temp_row,
                column=5
            ).value = (
                f"=INDEX(ITEM!D:D,"
                f"MATCH("
                f"INDEX('5'!B:B,"
                f"MATCH(A{temp_row},'5'!A:A,0)),"
                f"ITEM!A:A,0))"
            )

            # -----------------------------------------------------------------
            # Ledger Amount
            # -----------------------------------------------------------------

            temp5.cell(
                row=temp_row,
                column=6
            ).value = (
                f"=INDEX('5'!G:G,"
                f"MATCH(A{temp_row},'5'!A:A,0))"
            )

        # =====================================================================
        # COMMON FIELDS FOR CTR 1 TO LAST CTR
        # =====================================================================

        if ctr != 0:

            # -----------------------------------------------------------------
            # Item Name
            # -----------------------------------------------------------------

            temp5.cell(
                row=temp_row,
                column=8
            ).value = (
                f"=INDEX('5'!B:B,"
                f"MATCH(A{temp_row},'5'!A:A,0))"
            )

            # -----------------------------------------------------------------
            # Actual Quantity
            # -----------------------------------------------------------------

            temp5.cell(
                row=temp_row,
                column=10
            ).value = (
                f"=INDEX('5'!D:D,"
                f"MATCH(A{temp_row},'5'!A:A,0))"
            )

            # -----------------------------------------------------------------
            # Billed Quantity
            # -----------------------------------------------------------------

            temp5.cell(
                row=temp_row,
                column=11
            ).value = (
                f"=INDEX('5'!D:D,"
                f"MATCH(A{temp_row},'5'!A:A,0))"
            )

            # -----------------------------------------------------------------
            # Quantity UOM
            # -----------------------------------------------------------------

            temp5.cell(
                row=temp_row,
                column=12
            ).value = (
                f"=INDEX(ITEM!E:E,"
                f"MATCH("
                f"INDEX('5'!B:B,"
                f"MATCH(A{temp_row},'5'!A:A,0)),"
                f"ITEM!A:A,0))"
            )

            # -----------------------------------------------------------------
            # Item Rate
            # -----------------------------------------------------------------

            temp5.cell(
                row=temp_row,
                column=13
            ).value = (
                f"=INDEX('5'!F:F,"
                f"MATCH(A{temp_row},'5'!A:A,0))"
            )

            # -----------------------------------------------------------------
            # Item Rate Per
            # -----------------------------------------------------------------

            temp5.cell(
                row=temp_row,
                column=14
            ).value = (
                f"=INDEX(ITEM!E:E,"
                f"MATCH("
                f"INDEX('5'!B:B,"
                f"MATCH(A{temp_row},'5'!A:A,0)),"
                f"ITEM!A:A,0))"
            )

            # -----------------------------------------------------------------
            # Item Amount
            # -----------------------------------------------------------------

            temp5.cell(
                row=temp_row,
                column=15
            ).value = (
                f"=INDEX('5'!G:G,"
                f"MATCH(A{temp_row},'5'!A:A,0))"
            )

    logging.info(
        "Temp_5 formulas populated successfully"
    )

###############################################################################
# POPULATE TEMP_18
###############################################################################

def populate_temp_18(book, temp18, sheet18):

    logging.info("Populating Temp_18 formulas...")

    HEADER_ROW = book.header_row

    # -------------------------------------------------------------------------
    # Sheet 18 data rows
    # -------------------------------------------------------------------------

    first_sheet18_data_row = HEADER_ROW + 1

    last_sheet18_row = sheet18.max_row

    # -------------------------------------------------------------------------
    # Number of Sl rows in Sheet 18
    #
    # Example:
    #
    # Sl 1 to Sl 27 = 27 rows
    #
    # -------------------------------------------------------------------------

    total_sl_rows = last_sheet18_row - HEADER_ROW

    # -------------------------------------------------------------------------
    # Temp_18:
    #
    # Ctr 0 = Farmer Name
    # Ctr 1 = Sl 1
    # ...
    # Ctr N = Sl N
    #
    # -------------------------------------------------------------------------

    total_temp18_rows = total_sl_rows + 1

    # -------------------------------------------------------------------------
    # Final three rows in Sheet 18
    #
    # Sl N-2 = FREIGHT CHARGES FOR EC SHED
    # Sl N-1 = OUTPUT CGST @ 9%
    # Sl N   = OUTPUT SGST @ 9%
    #
    # -------------------------------------------------------------------------

    freight_ctr = total_sl_rows - 2

    cgst_ctr = total_sl_rows - 1

    sgst_ctr = total_sl_rows

    # -------------------------------------------------------------------------
    # Farmer ledger name
    # -------------------------------------------------------------------------

    farmer_ledger_name = get_farmer_ledger_name(book)

    # -------------------------------------------------------------------------
    # Populate rows
    # -------------------------------------------------------------------------

    for temp_row in range(2, total_temp18_rows + 2):

        # ---------------------------------------------------------------------
        # Ctr
        # ---------------------------------------------------------------------

        ctr = temp_row - 2

        # =====================================================================
        # CTR 0 — FARMER NAME
        # =====================================================================

        if ctr == 0:

            # Ledger Name
            temp18.cell(
                row=temp_row,
                column=5
            ).value = farmer_ledger_name

            # Ledger Amount
            #
            # Sum of all Final Amount values in Sheet 18
            #

            temp18.cell(
                row=temp_row,
                column=6
            ).value = (
                f"=SUM('18'!K{first_sheet18_data_row}:"
                f"K{last_sheet18_row})"
            )

        # =====================================================================
        # CGST SUMMARY ROW
        # =====================================================================

        elif ctr == cgst_ctr:

            # Ledger Name
            temp18.cell(
                row=temp_row,
                column=5
            ).value = (
                f"=INDEX('18'!B:B,"
                f"MATCH(A{temp_row},'18'!A:A,0))"
            )

            # Ledger Amount = CGST total
            temp18.cell(
                row=temp_row,
                column=6
            ).value = (
                f"=INDEX('18'!I:I,"
                f"MATCH(A{temp_row},'18'!A:A,0))"
            )

        # =====================================================================
        # SGST SUMMARY ROW
        # =====================================================================

        elif ctr == sgst_ctr:

            # Ledger Name
            temp18.cell(
                row=temp_row,
                column=5
            ).value = (
                f"=INDEX('18'!B:B,"
                f"MATCH(A{temp_row},'18'!A:A,0))"
            )

            # Ledger Amount = SGST total
            temp18.cell(
                row=temp_row,
                column=6
            ).value = (
                f"=INDEX('18'!J:J,"
                f"MATCH(A{temp_row},'18'!A:A,0))"
            )

# =====================================================================
# FREIGHT ROW
# =====================================================================
            #freight_ctr = total_sl_rows - 2
        elif ctr == freight_ctr:

    # Ledger Name
    #
    # Ctr 22 -> Sheet 18 Sl 22
    #
    # Freight ledger name should come directly from Sheet 18.
    # It should NOT be looked up in ITEM sheet.

            temp18.cell(
                row=temp_row,
                column=5
            ).value = (
                f"=INDEX('18'!B:B,"
                f"MATCH(A{temp_row},'18'!A:A,0))"
            )

    # Ledger Amount
    #
    # Freight amount comes from Item Value / Column G
    #

            temp18.cell(
                row=temp_row,
                column=6
            ).value = (
                f"=INDEX('18'!G:G,"
                f"MATCH(A{temp_row},'18'!A:A,0))"
            )
# =====================================================================
# NORMAL ITEM ROW
# =====================================================================        
        else:

    # Ledger Name
    #
    # Normal items are looked up in ITEM sheet.

            temp18.cell(
                row=temp_row,
                column=5
            ).value = (
                f"=INDEX(ITEM!D:D,"
                f"MATCH("
                f"INDEX('18'!B:B,"
                f"MATCH(A{temp_row},'18'!A:A,0)),"
                f"ITEM!A:A,0))"
            )

    # Ledger Amount

            temp18.cell(
                row=temp_row,
                column=6
            ).value = (
                f"=INDEX('18'!G:G,"
                f"MATCH(A{temp_row},'18'!A:A,0))"
            )


            # -----------------------------------------------------------------
            # Ledger Amount
            #
            # For all rows except the final two GST summary rows:
            #
            # Sheet 18 Column G
            #
            # This includes the Freight row.
            # -----------------------------------------------------------------

            temp18.cell(
                row=temp_row,
                column=6
            ).value = (
                f"=INDEX('18'!G:G,"
                f"MATCH(A{temp_row},'18'!A:A,0))"
            )

        # =====================================================================
        # COMMON FIELDS FOR CTR 1 ONWARDS
        # =====================================================================

        if ctr != 0:

            # -----------------------------------------------------------------
            # Item Name
            # -----------------------------------------------------------------

            temp18.cell(
                row=temp_row,
                column=8
            ).value = (
                f"=INDEX('18'!B:B,"
                f"MATCH(A{temp_row},'18'!A:A,0))"
            )

            # -----------------------------------------------------------------
            # Actual Quantity
            # -----------------------------------------------------------------

            temp18.cell(
                row=temp_row,
                column=10
            ).value = (
                f"=INDEX('18'!D:D,"
                f"MATCH(A{temp_row},'18'!A:A,0))"
            )

            # -----------------------------------------------------------------
            # Billed Quantity
            # -----------------------------------------------------------------

            temp18.cell(
                row=temp_row,
                column=11
            ).value = (
                f"=INDEX('18'!D:D,"
                f"MATCH(A{temp_row},'18'!A:A,0))"
            )

            # -----------------------------------------------------------------
            # Quantity UOM
            # -----------------------------------------------------------------

            temp18.cell(
                row=temp_row,
                column=12
            ).value = (
                f"=INDEX(ITEM!E:E,"
                f"MATCH("
                f"INDEX('18'!B:B,"
                f"MATCH(A{temp_row},'18'!A:A,0)),"
                f"ITEM!A:A,0))"
            )

            # -----------------------------------------------------------------
            # Item Rate
            # -----------------------------------------------------------------

            temp18.cell(
                row=temp_row,
                column=13
            ).value = (
                f"=INDEX('18'!F:F,"
                f"MATCH(A{temp_row},'18'!A:A,0))"
            )

            # -----------------------------------------------------------------
            # Item Rate Per
            # -----------------------------------------------------------------

            temp18.cell(
                row=temp_row,
                column=14
            ).value = (
                f"=INDEX(ITEM!E:E,"
                f"MATCH("
                f"INDEX('18'!B:B,"
                f"MATCH(A{temp_row},'18'!A:A,0)),"
                f"ITEM!A:A,0))"
            )

            # -----------------------------------------------------------------
            # Item Amount
            # -----------------------------------------------------------------

            temp18.cell(
                row=temp_row,
                column=15
            ).value = (
                f"=INDEX('18'!G:G,"
                f"MATCH(A{temp_row},'18'!A:A,0))"
            )

    logging.info(
        "Temp_18 formulas populated successfully"
    )

###############################################################################
# SELECT EXCEL FILE
###############################################################################

def select_excel_file():

    root = tk.Tk()

    # Hide the empty Tkinter window
    root.withdraw()

    # Keep the file dialog above other windows
    root.attributes("-topmost", True)

    filename = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[
            (
                "Excel Files",
                "*.xlsx *.xlsm *.xltx *.xltm"
            ),
            (
                "All Files",
                "*.*"
            )
        ]
    )

    # Close Tkinter after selecting the file
    root.destroy()

    return filename

###############################################################################
# PHASE-1
###############################################################################

def phase1(filename):

    logging.info("=" * 60)

    logging.info("PHASE-1 STARTED")

    logging.info("=" * 60)

    book = GSTWorkbook(filename)

    # ---------------------------------------------------------------
    # Create Sheets 5 and 18
    # ---------------------------------------------------------------

    sheet5, sheet18 = create_output_sheets(book)

    # ---------------------------------------------------------------
    # Copy header section
    # ---------------------------------------------------------------

    copy_header(
        book,
        sheet5
    )

    copy_header(
        book,
        sheet18
    )

    # ---------------------------------------------------------------
    # Split GST rows
    # ---------------------------------------------------------------

    split_gst(
        book,
        sheet5,
        sheet18
    )

    # ---------------------------------------------------------------
    # Calculate item values and GST
    # ---------------------------------------------------------------

    logging.info(
        "Calculating values in Sheet 5..."
    )

    calculate_bill_values(
        sheet5,
        book.header_row
    )

    logging.info(
        "Calculating values in Sheet 18..."
    )

    calculate_bill_values(
        sheet18,
        book.header_row
    )

    # ---------------------------------------------------------------
    # Add GST summary rows
    # ---------------------------------------------------------------

    complete_gst_sheet(
        sheet5,
        book.header_row,
        "OUTPUT CGST @ 2.5%",
        "OUTPUT SGST @ 2.5%"
    )

    complete_gst_sheet(
        sheet18,
        book.header_row,
        "OUTPUT CGST @ 9%",
        "OUTPUT SGST @ 9%"
    )

###############################################################################
# CREATE TEMP_5
###############################################################################

    temp5=create_temp_5(
        book,
        sheet5
)
    populate_temp_5(
    book,
    temp5,
    sheet5
)

###############################################################################
# CREATE AND POPULATE TEMP_18
###############################################################################

    temp18 = create_temp_18(
        book,
        sheet18
)

    populate_temp_18(
        book,
        temp18,
        sheet18
)

    # ---------------------------------------------------------------
    # Save final workbook
    # ---------------------------------------------------------------

    save(book)

    logging.info("=" * 60)

    logging.info(
        "PHASE-1 COMPLETED"
    )

    logging.info("=" * 60)


###############################################################################
# TEST
###############################################################################

#if __name__ == "__main__":

    #filename = input(
        #"\nEnter Excel File Path : "
    #).strip()

    #phase1(filename)
if __name__ == "__main__":

    print(
        "\n"
        + "=" * 70
    )

    print(
        "PART-1 LOADED SUCCESSFULLY"
    )

    print(
        "=" * 70
    )

    print(
        "\nPlease select the Excel file from the Windows file dialog..."
    )

    filename = select_excel_file()

    if not filename:

        print(
            "\nNo Excel file selected."
        )

        exit()

    print(
        f"\nSelected Excel File:"
    )

    print(
        filename
    )

    phase1(filename)